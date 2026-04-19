#!/usr/bin/env python3
"""
Generate a Project+ Work Breakdown Structure .xlsx from a JSON input.

Renders a hierarchical WBS table with code-based indentation, rolls up
effort/duration for parent elements, highlights TBDs, flags 100%-rule
and 8–80 hour rule violations on an Integrity sheet, and writes a
companion open_questions.md when gaps are large.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TBD_PATTERN = re.compile(r"\[TBD[^\]]*\]")

# Colors
TBD_FONT_COLOR = "B7471E"      # burnt orange
HEADER_FILL = "1F3A5F"         # deep navy
HEADER_FONT_COLOR = "FFFFFF"
LEVEL_FILLS = {
    1: "1F3A5F",   # deep navy — project
    2: "DCE6F1",   # light blue — phase / deliverable
    3: "F2F4F7",   # light gray — work package
}
LEVEL_FONT_COLORS = {
    1: "FFFFFF",
    2: "1F3A5F",
    3: "333333",
}
INTEGRITY_COLOR = "B7471E"


def is_tbd(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return bool(TBD_PATTERN.search(value)) or not value.strip()
    return False


def thin_border() -> Border:
    side = Side(border_style="thin", color="B0B7C3")
    return Border(left=side, right=side, top=side, bottom=side)


def to_float(value):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and not is_tbd(value):
        try:
            return float(value)
        except ValueError:
            return None
    return None


def code_depth(code: str) -> int:
    if not code:
        return 0
    return code.count(".") + 1


def parent_code(code: str) -> str:
    if "." not in code:
        return ""
    return code.rsplit(".", 1)[0]


def sort_code_key(code: str):
    parts = []
    for segment in code.split("."):
        try:
            parts.append((0, int(segment)))
        except ValueError:
            parts.append((1, segment))
    return tuple(parts)


def collect_tbds(data: dict) -> list[str]:
    gaps: list[str] = []
    header = data.get("header", {}) or {}
    for k, v in header.items():
        if is_tbd(v):
            gaps.append(f"header.{k}: {v or '[TBD — empty]'}")
    elements = data.get("elements") or []
    if not elements:
        gaps.append("elements: (empty — no WBS elements provided)")
    for i, el in enumerate(elements):
        label = el.get("code") or f"element[{i}]"
        for field in ("code", "level", "name", "type"):
            if is_tbd(el.get(field)):
                gaps.append(f"{label}.{field}: {el.get(field) or '[TBD — empty]'}")
        for field in ("description", "owner", "effort_hours", "duration_days"):
            if is_tbd(el.get(field)):
                gaps.append(f"{label}.{field}: {el.get(field)}")
    return gaps


def integrity_issues(elements: list[dict]) -> list[str]:
    """100% rule and 8–80 work package rule checks.

    Fully-TBD cells are not flagged here — those are gaps, not integrity
    violations. We only flag concrete values that violate the rules.
    """
    issues: list[str] = []
    codes = {el.get("code"): el for el in elements}

    # 100% rule: every non-project, non-work-package element should have children
    for el in elements:
        code = el.get("code", "")
        el_type = (el.get("type") or "").strip()
        if el_type in {"Phase", "Deliverable"}:
            has_child = any(
                other.get("code", "").startswith(code + ".")
                for other in elements
            )
            if not has_child:
                issues.append(
                    f"{code} '{el.get('name', '')}' is a {el_type} "
                    f"but has no child work packages (100% rule)."
                )

    # 8–80 hour rule on work packages
    for el in elements:
        if (el.get("type") or "").strip() != "Work Package":
            continue
        eff = to_float(el.get("effort_hours"))
        if eff is None:
            continue  # TBD — not a violation
        if eff < 8:
            issues.append(
                f"{el.get('code')} '{el.get('name')}' has {eff:g}h effort — "
                "below 8-hour work package floor (consider combining)."
            )
        elif eff > 80:
            issues.append(
                f"{el.get('code')} '{el.get('name')}' has {eff:g}h effort — "
                "above 80-hour work package ceiling (needs further decomposition)."
            )

    # Dependency validation
    for el in elements:
        deps = el.get("dependencies", "")
        if not deps or is_tbd(str(deps)):
            continue
        for dep in str(deps).split(","):
            dep = dep.strip()
            if dep and dep not in codes:
                issues.append(
                    f"{el.get('code')} '{el.get('name')}' depends on "
                    f"'{dep}' which does not exist in the WBS."
                )

    return issues


def rollup(elements: list[dict]) -> dict[str, dict]:
    """Roll up effort/duration for non-work-package elements from their
    descendants. Only concrete numbers contribute — TBDs are ignored in the
    rollup total and surfaced separately as a "partial" flag.
    """
    rollups: dict[str, dict] = {}
    for el in elements:
        code = el.get("code", "")
        el_type = (el.get("type") or "").strip()
        if el_type == "Work Package":
            continue
        total_effort = 0.0
        total_duration = 0.0
        has_effort = False
        has_duration = False
        partial = False
        for other in elements:
            other_code = other.get("code", "")
            if not other_code.startswith(code + "."):
                continue
            if (other.get("type") or "").strip() != "Work Package":
                continue
            eff = to_float(other.get("effort_hours"))
            dur = to_float(other.get("duration_days"))
            if eff is not None:
                total_effort += eff
                has_effort = True
            elif is_tbd(other.get("effort_hours")):
                partial = True
            if dur is not None:
                total_duration += dur
                has_duration = True
            elif is_tbd(other.get("duration_days")):
                partial = True
        rollups[code] = {
            "effort": total_effort if has_effort else None,
            "duration": total_duration if has_duration else None,
            "partial": partial,
        }
    return rollups


def write_wbs_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "WBS"

    header = data.get("header", {}) or {}
    project_name = header.get("project_name", "[TBD — no project name]")
    elements = list(data.get("elements") or [])
    elements.sort(key=lambda el: sort_code_key(el.get("code", "")))

    cols = [
        ("code", "Code", 10),
        ("name", "Name", 44),
        ("type", "Type", 14),
        ("description", "Description", 38),
        ("owner", "Owner", 20),
        ("effort_hours", "Effort (h)", 12),
        ("duration_days", "Duration (d)", 12),
        ("dependencies", "Dependencies", 16),
    ]
    ncols = len(cols)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=f"{project_name} — Work Breakdown Structure")
    tc.font = Font(bold=True, size=16, color=HEADER_FILL)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Meta
    meta = [
        ("Project Manager", header.get("project_manager", "[TBD]")),
        ("Sponsor", header.get("sponsor", "[TBD]")),
        ("Date", header.get("date", "[TBD]")),
        ("Version", header.get("version", "v1")),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=ncols)
        vc = ws.cell(row=i, column=2, value=str(v))
        if is_tbd(str(v)):
            vc.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
        else:
            vc.font = Font(size=11)

    # Legend
    legend_row = 2 + len(meta) + 1
    legend = ws.cell(
        row=legend_row, column=1,
        value="Legend: rows indented by WBS code depth. "
              "Non-leaf rows show rolled-up effort/duration from children.",
    )
    legend.font = Font(italic=True, size=10, color="555555")
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=ncols)

    # Column headers
    header_row = legend_row + 2
    for c_idx, (_, label, _) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[header_row].height = 22

    # Rollups
    rollups = rollup(elements)

    # Rows
    if not elements:
        nr = header_row + 1
        ws.cell(row=nr, column=1, value="[TBD — no WBS elements provided]").font = Font(
            color=TBD_FONT_COLOR, italic=True
        )
    else:
        for r_offset, el in enumerate(elements, start=1):
            row = header_row + r_offset
            level = el.get("level") or code_depth(el.get("code", ""))
            try:
                level = int(level)
            except (ValueError, TypeError):
                level = 1
            level_clamped = max(1, min(level, 3))
            fill_color = LEVEL_FILLS.get(level_clamped, "FFFFFF")
            font_color = LEVEL_FONT_COLORS.get(level_clamped, "333333")
            bold = level_clamped <= 2

            el_type = (el.get("type") or "").strip()
            code = el.get("code", "")
            rolled = rollups.get(code) if el_type != "Work Package" else None

            # Values
            name_indented = ("    " * max(0, level - 1)) + str(el.get("name", "[TBD]"))
            values = [
                el.get("code", "[TBD]"),
                name_indented,
                el_type or "[TBD]",
                el.get("description", "[TBD]"),
                el.get("owner", "[TBD]"),
            ]
            # Effort column
            if el_type == "Work Package":
                values.append(el.get("effort_hours", "[TBD]"))
                values.append(el.get("duration_days", "[TBD]"))
            else:
                if rolled and rolled["effort"] is not None:
                    effort_str = f"{rolled['effort']:g}"
                    if rolled["partial"]:
                        effort_str += "*"
                    values.append(effort_str)
                else:
                    values.append("[TBD]" if rolled and rolled["partial"] else "—")
                if rolled and rolled["duration"] is not None:
                    dur_str = f"{rolled['duration']:g}"
                    if rolled["partial"]:
                        dur_str += "*"
                    values.append(dur_str)
                else:
                    values.append("[TBD]" if rolled and rolled["partial"] else "—")
            values.append(el.get("dependencies", ""))

            for c_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c_idx, value=str(v) if v is not None else "")
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = thin_border()
                if is_tbd(str(v)):
                    cell.font = Font(color=TBD_FONT_COLOR, italic=True,
                                     bold=bold, size=11)
                else:
                    cell.font = Font(color=font_color, bold=bold, size=11)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for c_idx, (_, _, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    # Footnote for partial rollups
    if any(r and r.get("partial") for r in rollups.values()):
        note_row = header_row + len(elements) + 2
        note = ws.cell(row=note_row, column=1, value=(
            "* = partial rollup; some children have TBD effort/duration. "
            "Resolve the TBDs on the Draft Gaps sheet to get complete totals."
        ))
        note.font = Font(italic=True, size=10, color="555555")
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=ncols)


def write_summary_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Work Package Summary")
    elements = data.get("elements") or []
    wps = [el for el in elements if (el.get("type") or "").strip() == "Work Package"]

    ws.cell(row=1, column=1, value="Work Package Summary").font = Font(
        bold=True, size=14, color=HEADER_FILL
    )
    intro = ws.cell(row=2, column=1, value=(
        "All work packages (lowest-level WBS elements). "
        "Use this as the starting point for schedule and budget."
    ))
    intro.font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    cols = ["Code", "Name", "Owner", "Effort (h)", "Duration (d)"]
    for i, label in enumerate(cols, start=1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

    for r, wp in enumerate(wps, start=5):
        row_vals = [
            wp.get("code", "[TBD]"),
            wp.get("name", "[TBD]"),
            wp.get("owner", "[TBD]"),
            wp.get("effort_hours", "[TBD]"),
            wp.get("duration_days", "[TBD]"),
        ]
        for i, v in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=i, value=str(v))
            c.border = thin_border()
            if is_tbd(str(v)):
                c.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
            else:
                c.font = Font(size=11)

    # Totals row
    total_effort = sum(
        to_float(wp.get("effort_hours")) or 0 for wp in wps
        if to_float(wp.get("effort_hours")) is not None
    )
    total_duration = sum(
        to_float(wp.get("duration_days")) or 0 for wp in wps
        if to_float(wp.get("duration_days")) is not None
    )
    total_row = 5 + len(wps) + 1
    tc = ws.cell(row=total_row, column=1, value="TOTAL (concrete only)")
    tc.font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=4, value=f"{total_effort:g}").font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=5, value=f"{total_duration:g}").font = Font(bold=True, size=11)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14


def write_gaps_sheet(wb: Workbook, gaps: list[str]) -> None:
    ws = wb.create_sheet("Draft Gaps")
    ws.cell(row=1, column=1, value=f"Draft Gaps ({len(gaps)})").font = Font(
        bold=True, size=14, color=HEADER_FILL
    )
    intro = ws.cell(row=2, column=1, value=(
        "Unresolved items in this WBS draft. Every TBD is a potential "
        "source of schedule or budget surprises — resolve with the sponsor."
    ))
    intro.font = Font(italic=True, size=10)
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.row_dimensions[2].height = 42
    for i, g in enumerate(gaps, start=4):
        c = ws.cell(row=i, column=1, value=g)
        c.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    ws.column_dimensions["A"].width = 80


def write_integrity_sheet(wb: Workbook, issues: list[str]) -> None:
    ws = wb.create_sheet("Integrity Issues")
    ws.cell(row=1, column=1, value=f"WBS Integrity Issues ({len(issues)})").font = Font(
        bold=True, size=14, color=INTEGRITY_COLOR
    )
    intro = ws.cell(row=2, column=1, value=(
        "Project+ WBS rules: deliverables/phases need child work packages "
        "(100% rule); work packages must be 8–80 hours; dependencies must "
        "reference valid codes. These violations should be fixed."
    ))
    intro.font = Font(italic=True, size=10)
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.row_dimensions[2].height = 42
    for i, issue in enumerate(issues, start=4):
        c = ws.cell(row=i, column=1, value=issue)
        c.font = Font(color=INTEGRITY_COLOR, size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    ws.column_dimensions["A"].width = 95


def render_wbs(data: dict, output_path: Path) -> tuple[list[str], list[str]]:
    wb = Workbook()
    write_wbs_sheet(wb, data)
    write_summary_sheet(wb, data)
    gaps = collect_tbds(data)
    issues = integrity_issues(data.get("elements") or [])
    if gaps:
        write_gaps_sheet(wb, gaps)
    if issues:
        write_integrity_sheet(wb, issues)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return gaps, issues


def write_open_questions(gaps: list[str], issues: list[str], output_path: Path) -> None:
    lines = [
        "# Open Questions for Sponsor / PM",
        "",
        f"This WBS draft has **{len(gaps)} unresolved items**"
        + (f" and **{len(issues)} integrity issues**" if issues else "")
        + ". Resolve these before using the WBS for scheduling/budgeting.",
        "",
        "## Draft Gaps",
        "",
    ]
    for g in gaps:
        lines.append(f"- {g}")
    if issues:
        lines += ["", "## Integrity Issues (Project+ rule violations)", ""]
        for i in issues:
            lines.append(f"- {i}")
    lines.append("")
    output_path.write_text("\n".join(lines))


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a Project+ WBS .xlsx")
    parser.add_argument("--input", help="Path to JSON input (omit to read stdin)")
    parser.add_argument("--output", required=True)
    parser.add_argument("--questions-threshold", type=int, default=5)
    args = parser.parse_args()

    output_path = Path(args.output)
    try:
        if args.input:
            data = json.loads(Path(args.input).read_text())
        else:
            data = json.loads(sys.stdin.read())
    except Exception as e:
        print(f"ERROR: could not parse input JSON: {e}", file=sys.stderr)
        return 2

    gaps, issues = render_wbs(data, output_path)
    print(f"WBS written: {output_path}")
    print(f"TBD count: {len(gaps)}")
    if issues:
        print(f"Integrity issues: {len(issues)}")

    if len(gaps) >= args.questions_threshold or issues:
        oq = output_path.parent / "open_questions.md"
        write_open_questions(gaps, issues, oq)
        print(f"Open questions written: {oq}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
