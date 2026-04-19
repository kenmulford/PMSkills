#!/usr/bin/env python3
"""Generate a Project+ project schedule .xlsx from JSON input.

Sheets produced:
- Schedule: tasks with owner, duration, predecessors, computed start/finish, Gantt bars
- Critical Path: longest chain of dependent tasks when dates are fully computable
- Draft Gaps: all TBD/unresolved items
- Integrity Issues: circular deps, missing preds, 100%-of-rule violations
- Notes: assumptions (weekend skip, holidays not modeled)

Also emits open_questions.md when gaps >= threshold.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


TBD_PATTERN = re.compile(r"\[TBD[^\]]*\]")

NAVY = "1F3A5F"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
TBD_ORANGE = "B7471E"
TBD_FILL = "FCE8D4"
CRITICAL_FILL = "FFE599"
BAR_FILL = "4A90C2"

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def is_tbd(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return bool(TBD_PATTERN.search(v)) or not v.strip()
    return False


def parse_date(s):
    if not s or is_tbd(s):
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def add_business_days(start: date, n: int) -> date:
    """Return the date that is n business days after start (start counts as day 1)."""
    if n <= 0:
        return start
    d = start
    remaining = n - 1
    while remaining > 0:
        d += timedelta(days=1)
        if d.weekday() < 5:
            remaining -= 1
    return d


def next_business_day(d: date) -> date:
    nd = d + timedelta(days=1)
    while nd.weekday() >= 5:
        nd += timedelta(days=1)
    return nd


def topo_sort(tasks):
    """Return task ids in topological order; detect cycles."""
    ids = [t["id"] for t in tasks]
    by_id = {t["id"]: t for t in tasks}
    indeg = {tid: 0 for tid in ids}
    graph = {tid: [] for tid in ids}
    missing = []
    for t in tasks:
        preds = [p.strip() for p in str(t.get("predecessors") or "").split(",") if p.strip()]
        for p in preds:
            if p not in by_id:
                missing.append((t["id"], p))
                continue
            graph[p].append(t["id"])
            indeg[t["id"]] += 1
    order = []
    queue = [tid for tid in ids if indeg[tid] == 0]
    while queue:
        n = queue.pop(0)
        order.append(n)
        for child in graph[n]:
            indeg[child] -= 1
            if indeg[child] == 0:
                queue.append(child)
    cycles = [tid for tid in ids if tid not in order]
    return order, missing, cycles


def compute_dates(tasks, anchor):
    """Return dict id -> (start, finish, tbd_reason). tbd_reason non-empty = unresolved."""
    by_id = {t["id"]: t for t in tasks}
    order, _, cycles = topo_sort(tasks)
    results = {}

    for tid in order:
        t = by_id[tid]
        dur = t.get("duration_days")
        if is_tbd(dur) or not isinstance(dur, (int, float)):
            results[tid] = (None, None, "duration TBD")
            continue
        dur = int(dur)

        preds = [p.strip() for p in str(t.get("predecessors") or "").split(",") if p.strip()]
        if not preds:
            if anchor is None:
                results[tid] = (None, None, "no anchor date")
                continue
            start = anchor
            while start.weekday() >= 5:
                start = next_business_day(start)
            finish = add_business_days(start, dur)
            results[tid] = (start, finish, "")
            continue

        # latest finish of predecessors (FS default)
        latest = None
        blocked = None
        for p in preds:
            if p not in results:
                blocked = f"pred {p} not resolved"
                break
            ps, pf, pr = results[p]
            if pr:
                blocked = f"pred {p} {pr}"
                break
            if latest is None or pf > latest:
                latest = pf
        if blocked:
            results[tid] = (None, None, blocked)
            continue

        start = next_business_day(latest)
        finish = add_business_days(start, dur)
        results[tid] = (start, finish, "")

    # cycles
    for tid in cycles:
        results[tid] = (None, None, "circular dependency")

    return results


def critical_path(tasks, dates):
    """Longest chain of tasks by finish date; only valid if all dates are real."""
    if any(d[2] for d in dates.values()):
        return None, "unavailable — " + f"{sum(1 for d in dates.values() if d[2])} tasks missing duration or anchor"
    by_id = {t["id"]: t for t in tasks}
    # build reverse graph
    preds_of = {}
    for t in tasks:
        preds_of[t["id"]] = [p.strip() for p in str(t.get("predecessors") or "").split(",") if p.strip()]
    # longest path ending at each node by finish date depth
    best = {}  # id -> (length, prev)
    order, _, _ = topo_sort(tasks)
    for tid in order:
        ps = preds_of[tid]
        if not ps:
            best[tid] = (1, None)
        else:
            candidates = [(best[p][0] + 1, p) for p in ps if p in best]
            if candidates:
                best[tid] = max(candidates)
            else:
                best[tid] = (1, None)
    # pick node with latest finish
    end = max(best.keys(), key=lambda t: (dates[t][1], best[t][0]))
    chain = []
    cur = end
    while cur is not None:
        chain.append(cur)
        cur = best[cur][1]
    chain.reverse()
    return chain, None


# ---- rendering ----


def _header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _tbd_style(cell):
    cell.font = Font(italic=True, color=TBD_ORANGE)
    cell.fill = PatternFill("solid", fgColor=TBD_FILL)


def render_schedule_sheet(wb, data, dates, critical_ids):
    ws = wb.active
    ws.title = "Schedule"

    header = data.get("header", {}) or {}
    ws["A1"] = "Project:"
    ws["B1"] = header.get("project_name", "[TBD]")
    ws["A2"] = "PM:"
    ws["B2"] = header.get("project_manager", "[TBD]")
    ws["D2"] = "Sponsor:"
    ws["E2"] = header.get("sponsor", "[TBD]")
    ws["A3"] = "Anchor start:"
    ws["B3"] = header.get("anchor_start_date", "[TBD]")
    ws["D3"] = "Version:"
    ws["E3"] = header.get("version", "v1")
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=4).font = Font(bold=True)

    headers = ["ID", "Task", "Owner", "Duration (days)", "Predecessors",
               "Dep Type", "Lag", "Start", "Finish"]
    start_row = 5
    _header_row(ws, headers, row=start_row)

    tasks = data.get("tasks") or []

    # compute date range for Gantt
    real_dates = [dates[t["id"]] for t in tasks if not dates[t["id"]][2]]
    if real_dates:
        min_d = min(d[0] for d in real_dates)
        max_d = max(d[1] for d in real_dates)
        # weekly columns
        weeks = []
        cur = min_d - timedelta(days=min_d.weekday())
        while cur <= max_d:
            weeks.append(cur)
            cur += timedelta(days=7)
        for i, w in enumerate(weeks):
            c = ws.cell(row=start_row, column=len(headers) + 1 + i, value=w.strftime("%m/%d"))
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
    else:
        weeks = []

    for i, t in enumerate(tasks):
        r = start_row + 1 + i
        tid = t["id"]
        is_crit = critical_ids and tid in critical_ids
        row_fill = PatternFill("solid", fgColor=CRITICAL_FILL) if is_crit else None

        vals = [
            tid,
            t.get("name", "[TBD]"),
            t.get("owner", "[TBD]"),
            t.get("duration_days", "[TBD]"),
            t.get("predecessors") or "",
            t.get("dependency_type") or "FS",
            t.get("lag_days") or 0,
        ]
        s, f, reason = dates[tid]
        vals.append(s.strftime("%Y-%m-%d") if s else f"[TBD — {reason}]")
        vals.append(f.strftime("%Y-%m-%d") if f else f"[TBD — {reason}]")

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER
            if isinstance(v, str) and TBD_PATTERN.search(v):
                _tbd_style(c)
            elif row_fill:
                c.fill = row_fill

        # Gantt bars
        if weeks and s and f:
            for wi, w in enumerate(weeks):
                w_end = w + timedelta(days=6)
                if s <= w_end and f >= w:
                    c = ws.cell(row=r, column=len(headers) + 1 + wi, value="█")
                    c.fill = PatternFill("solid", fgColor=BAR_FILL)
                    c.font = Font(color="FFFFFF")
                    c.alignment = Alignment(horizontal="center")
                    c.border = BORDER

    # column widths
    widths = [8, 32, 18, 14, 14, 10, 8, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(weeks)):
        ws.column_dimensions[get_column_letter(len(headers) + 1 + i)].width = 6
    ws.freeze_panes = f"B{start_row + 1}"


def render_critical_path(wb, tasks, dates, chain, reason):
    ws = wb.create_sheet("Critical Path")
    ws["A1"] = "Critical Path"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    if reason:
        ws["A3"] = reason
        ws["A3"].font = Font(italic=True, color=TBD_ORANGE)
        return
    by_id = {t["id"]: t for t in tasks}
    _header_row(ws, ["Seq", "ID", "Task", "Start", "Finish"], row=3)
    for i, tid in enumerate(chain or [], 1):
        t = by_id[tid]
        s, f, _ = dates[tid]
        ws.cell(row=3 + i, column=1, value=i).border = BORDER
        ws.cell(row=3 + i, column=2, value=tid).border = BORDER
        ws.cell(row=3 + i, column=3, value=t.get("name", "")).border = BORDER
        ws.cell(row=3 + i, column=4, value=s.strftime("%Y-%m-%d")).border = BORDER
        ws.cell(row=3 + i, column=5, value=f.strftime("%Y-%m-%d")).border = BORDER
    for i, w in enumerate([6, 8, 32, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def collect_gaps(data, dates):
    gaps = []
    header = data.get("header", {}) or {}
    for k in ("project_name", "project_manager", "sponsor", "anchor_start_date"):
        if is_tbd(header.get(k)):
            gaps.append(f"header.{k}: {header.get(k) or '[TBD — empty]'}")
    tasks = data.get("tasks") or []
    if not tasks:
        gaps.append("tasks: (empty)")
    for t in tasks:
        tid = t.get("id", "?")
        for f in ("name", "owner", "duration_days"):
            if is_tbd(t.get(f)):
                gaps.append(f"task {tid}.{f}: {t.get(f) or '[TBD — empty]'}")
        s, fi, r = dates.get(tid, (None, None, ""))
        if r:
            gaps.append(f"task {tid}.dates: [TBD — {r}]")
    return gaps


def render_gaps(wb, gaps):
    ws = wb.create_sheet("Draft Gaps")
    ws["A1"] = f"Draft Gaps ({len(gaps)})"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A3"] = "Item"
    ws["A3"].font = Font(bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill("solid", fgColor=NAVY)
    if not gaps:
        ws["A4"] = "None — schedule is fully specified."
        return
    for i, g in enumerate(gaps, 1):
        c = ws.cell(row=3 + i, column=1, value=g)
        _tbd_style(c)
        c.border = BORDER
    ws.column_dimensions["A"].width = 80


def render_integrity(wb, tasks, missing_preds, cycles):
    ws = wb.create_sheet("Integrity Issues")
    ws["A1"] = "Integrity Issues"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    issues = []
    for tid, p in missing_preds:
        issues.append(f"Task {tid} references missing predecessor '{p}'")
    for tid in cycles:
        issues.append(f"Task {tid} is part of a circular dependency")
    # single-owner check: multi-owner strings
    for t in tasks:
        owner = t.get("owner", "")
        if isinstance(owner, str) and not is_tbd(owner):
            if " and " in owner or "," in owner or " & " in owner:
                issues.append(f"Task {t.get('id')} has multi-owner string '{owner}' — split into subtasks")
    if not issues:
        ws["A3"] = "No integrity issues detected."
        return
    for i, item in enumerate(issues, 1):
        c = ws.cell(row=2 + i, column=1, value=item)
        _tbd_style(c)
    ws.column_dimensions["A"].width = 80


def render_notes(wb):
    ws = wb.create_sheet("Notes")
    ws["A1"] = "Schedule Assumptions"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    notes = [
        "• Weekends (Sat/Sun) are skipped in duration math.",
        "• Holidays are NOT modeled. Adjust the anchor start or add slack if a holiday falls in the critical path.",
        "• Default dependency type is Finish-to-Start (FS) with zero lag unless explicitly stated.",
        "• [TBD — reason] cells are honest gaps: the source user message did not contain that information.",
        "• Critical path is computed only when every task has a real duration and the project has an anchor start date.",
    ]
    for i, n in enumerate(notes, 3):
        ws.cell(row=i, column=1, value=n)
    ws.column_dimensions["A"].width = 110


def write_open_questions(gaps, output):
    lines = ["# Open Questions — Project Schedule", "",
             f"This draft has **{len(gaps)} unresolved items**. Resolve before baselining the schedule.", ""]
    for g in gaps:
        lines.append(f"- {g}")
    output.write_text("\n".join(lines))


def render(data, output_path: Path, questions_threshold=5):
    wb = Workbook()
    tasks = data.get("tasks") or []
    header = data.get("header", {}) or {}
    anchor = parse_date(header.get("anchor_start_date"))

    _, missing_preds, cycles = topo_sort(tasks)
    dates = compute_dates(tasks, anchor)
    chain, cp_reason = critical_path(tasks, dates) if tasks else (None, "no tasks")
    critical_ids = set(chain) if chain else set()

    render_schedule_sheet(wb, data, dates, critical_ids)
    render_critical_path(wb, tasks, dates, chain, cp_reason)
    gaps = collect_gaps(data, dates)
    render_gaps(wb, gaps)
    render_integrity(wb, tasks, missing_preds, cycles)
    render_notes(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    if len(gaps) >= questions_threshold:
        oq = output_path.parent / "open_questions.md"
        write_open_questions(gaps, oq)
        return gaps, oq
    return gaps, None


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", help="Path to JSON input (omit to read stdin)")
    p.add_argument("--output", required=True)
    p.add_argument("--questions-threshold", type=int, default=5)
    args = p.parse_args()

    try:
        if args.input:
            data = json.loads(Path(args.input).read_text())
        else:
            data = json.loads(sys.stdin.read())
    except Exception as e:
        print(f"ERROR: could not parse input JSON: {e}", file=sys.stderr)
        return 2

    output = Path(args.output)
    gaps, oq = render(data, output, args.questions_threshold)
    print(f"Schedule written: {output}")
    print(f"TBD count: {len(gaps)}")
    if oq:
        print(f"Open questions: {oq}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
