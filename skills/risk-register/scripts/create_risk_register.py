#!/usr/bin/env python3
"""Generate a Project+ risk register .xlsx from JSON input.

Sheets:
- Risk Register: main table with heat-mapped severity
- Heat Map: 3x4 matrix with risk IDs in each cell
- Draft Gaps: all TBD/unresolved items
- Integrity Issues: invalid values, reporting-line manager traps
- Notes: scoring matrix legend
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


TBD_PATTERN = re.compile(r"\[TBD[^\]]*\]")

NAVY = "1F3A5F"
LIGHT_GRAY = "F2F2F2"
TBD_ORANGE = "B7471E"
TBD_FILL = "FCE8D4"

SEV_COLORS = {
    "Critical": "C00000",
    "High": "E97132",
    "Medium": "FFC000",
    "Low": "70AD47",
}

PROBS = ["Low", "Medium", "High"]
IMPACTS = ["Low", "Medium", "High", "Critical"]

MATRIX = {
    ("High", "Low"): "Medium",
    ("High", "Medium"): "High",
    ("High", "High"): "Critical",
    ("High", "Critical"): "Critical",
    ("Medium", "Low"): "Low",
    ("Medium", "Medium"): "Medium",
    ("Medium", "High"): "High",
    ("Medium", "Critical"): "Critical",
    ("Low", "Low"): "Low",
    ("Low", "Medium"): "Low",
    ("Low", "High"): "Medium",
    ("Low", "Critical"): "High",
}

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def is_tbd(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return bool(TBD_PATTERN.search(v)) or not v.strip()
    return False


def compute_severity(p, i):
    if is_tbd(p) or is_tbd(i):
        return "[TBD — severity awaits P/I]"
    key = (p, i)
    return MATRIX.get(key, f"[TBD — non-standard P/I: {p}/{i}]")


def _header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _tbd_style(cell):
    cell.font = Font(italic=True, color=TBD_ORANGE)
    cell.fill = PatternFill("solid", fgColor=TBD_FILL)


def render_register(wb, data):
    ws = wb.active
    ws.title = "Risk Register"

    h = data.get("header", {}) or {}
    ws["A1"] = "Project:"
    ws["B1"] = h.get("project_name", "[TBD]")
    ws["A2"] = "PM:"
    ws["B2"] = h.get("project_manager", "[TBD]")
    ws["D2"] = "Sponsor:"
    ws["E2"] = h.get("sponsor", "[TBD]")
    ws["A3"] = "Date:"
    ws["B3"] = h.get("date", "[TBD]")
    ws["D3"] = "Version:"
    ws["E3"] = h.get("version", "v1")
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=4).font = Font(bold=True)

    headers = ["ID", "Risk", "Category", "Probability", "Impact", "Severity",
               "Owner", "Mitigation", "Contingency", "Status", "Trigger"]
    start_row = 5
    _header_row(ws, headers, row=start_row)

    risks = data.get("risks") or []
    for i, r in enumerate(risks):
        row = start_row + 1 + i
        severity = compute_severity(r.get("probability"), r.get("impact"))
        vals = [
            r.get("id", f"R{i+1:03d}"),
            r.get("risk", "[TBD]"),
            r.get("category", "[TBD]"),
            r.get("probability", "[TBD]"),
            r.get("impact", "[TBD]"),
            severity,
            r.get("owner", "[TBD]"),
            r.get("mitigation", "[TBD]"),
            r.get("contingency") or "",
            r.get("status") or "Open",
            r.get("trigger") or "",
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if isinstance(v, str) and TBD_PATTERN.search(v):
                _tbd_style(c)
            # heat-map severity cell
            if col == 6 and severity in SEV_COLORS:
                c.fill = PatternFill("solid", fgColor=SEV_COLORS[severity])
                c.font = Font(bold=True, color="FFFFFF")

    widths = [8, 40, 14, 12, 12, 12, 16, 40, 30, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"B{start_row + 1}"


def render_heatmap(wb, risks):
    ws = wb.create_sheet("Heat Map")
    ws["A1"] = "Probability × Impact Heat Map"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)

    # header row
    ws["B3"] = "Low"
    ws["C3"] = "Medium"
    ws["D3"] = "High"
    ws["E3"] = "Critical"
    for col in range(2, 6):
        c = ws.cell(row=3, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    ws["A3"] = "P \\ I"
    ws["A3"].font = Font(bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A3"].border = BORDER

    row_order = ["High", "Medium", "Low"]
    for ri, p in enumerate(row_order, 4):
        ws.cell(row=ri, column=1, value=p)
        ws.cell(row=ri, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=ri, column=1).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=ri, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=ri, column=1).border = BORDER
        for ci, imp in enumerate(IMPACTS, 2):
            sev = MATRIX[(p, imp)]
            ids = [r.get("id", "") for r in risks
                   if r.get("probability") == p and r.get("impact") == imp]
            c = ws.cell(row=ri, column=ci, value=", ".join(ids) if ids else "")
            c.fill = PatternFill("solid", fgColor=SEV_COLORS[sev])
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
            ws.row_dimensions[ri].height = 30

    for col, w in enumerate([12, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # unscored risks
    unscored = [r.get("id", "") for r in risks
                if is_tbd(r.get("probability")) or is_tbd(r.get("impact"))]
    if unscored:
        ws["A9"] = f"Unscored (TBD P or I): {', '.join(unscored)}"
        ws["A9"].font = Font(italic=True, color=TBD_ORANGE)


def collect_gaps(data):
    gaps = []
    h = data.get("header", {}) or {}
    for k in ("project_name", "project_manager", "sponsor"):
        if is_tbd(h.get(k)):
            gaps.append(f"header.{k}: {h.get(k) or '[TBD — empty]'}")
    risks = data.get("risks") or []
    if not risks:
        gaps.append("risks: (empty)")
    for r in risks:
        rid = r.get("id", "?")
        for f in ("risk", "probability", "impact", "owner", "mitigation"):
            if is_tbd(r.get(f)):
                gaps.append(f"{rid}.{f}: {r.get(f) or '[TBD — empty]'}")
    return gaps


def render_gaps(wb, gaps):
    ws = wb.create_sheet("Draft Gaps")
    ws["A1"] = f"Draft Gaps ({len(gaps)})"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    if not gaps:
        ws["A3"] = "None — register is fully specified."
        return
    for i, g in enumerate(gaps, 1):
        c = ws.cell(row=2 + i, column=1, value=g)
        _tbd_style(c)
        c.border = BORDER
    ws.column_dimensions["A"].width = 80


def render_integrity(wb, risks):
    ws = wb.create_sheet("Integrity Issues")
    ws["A1"] = "Integrity Issues"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)

    issues = []
    for r in risks:
        rid = r.get("id", "?")
        p = r.get("probability")
        if not is_tbd(p) and p not in PROBS:
            issues.append(f"{rid}: invalid probability '{p}' — must be Low/Medium/High or [TBD]")
        i = r.get("impact")
        if not is_tbd(i) and i not in IMPACTS:
            issues.append(f"{rid}: invalid impact '{i}' — must be Low/Medium/High/Critical or [TBD]")
        status = r.get("status") or "Open"
        if status not in ("Open", "In Progress", "Closed", "Unconfirmed"):
            issues.append(f"{rid}: non-standard status '{status}'")

    if not issues:
        ws["A3"] = "No integrity issues detected."
        return
    for i, item in enumerate(issues, 1):
        c = ws.cell(row=2 + i, column=1, value=item)
        _tbd_style(c)
    ws.column_dimensions["A"].width = 100


def render_notes(wb):
    ws = wb.create_sheet("Notes")
    ws["A1"] = "Risk Register Conventions"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    notes = [
        "• Probability: Low / Medium / High.",
        "• Impact: Low / Medium / High / Critical.",
        "• Severity matrix (3×4) — see Heat Map sheet for the full grid.",
        "• Status: Open / In Progress / Closed / Unconfirmed.",
        "• Unconfirmed = user flagged the risk but has not validated probability or impact.",
        "• [TBD — reason] cells are honest gaps: the source user message did not contain that information.",
        "• Response strategies (Project+): avoid, transfer, mitigate, accept.",
    ]
    for i, n in enumerate(notes, 3):
        ws.cell(row=i, column=1, value=n)
    ws.column_dimensions["A"].width = 110


def write_open_questions(gaps, output):
    lines = ["# Open Questions — Risk Register", "",
             f"This draft has **{len(gaps)} unresolved items**. Resolve before reviewing with the sponsor.", ""]
    for g in gaps:
        lines.append(f"- {g}")
    output.write_text("\n".join(lines))


def render(data, output_path: Path, questions_threshold=5):
    wb = Workbook()
    render_register(wb, data)
    render_heatmap(wb, data.get("risks") or [])
    gaps = collect_gaps(data)
    render_gaps(wb, gaps)
    render_integrity(wb, data.get("risks") or [])
    render_notes(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    if len(gaps) >= questions_threshold:
        oq = output_path.parent / "open_questions.md"
        write_open_questions(gaps, oq)
        return gaps, oq
    return gaps, None


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", help="Path to JSON input (omit to read stdin)")
    p.add_argument("--output", required=True)
    p.add_argument("--questions-threshold", type=int, default=5)
    args = p.parse_args()

    try:
        if args.input:
            data = json.loads(Path(args.input).read_text())
        else:
            data = json.loads(sys.stdin.read())
    except Exception as e:
        print(f"ERROR: could not parse input JSON: {e}", file=sys.stderr)
        return 2

    out = Path(args.output)
    gaps, oq = render(data, out, args.questions_threshold)
    print(f"Risk register written: {out}")
    print(f"TBD count: {len(gaps)}")
    if oq:
        print(f"Open questions: {oq}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
