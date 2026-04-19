#!/usr/bin/env python3
"""Generate a project decision log .xlsx from JSON input.

Refuses to invent decision makers, rationale, or options considered.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TBD_PATTERN = re.compile(r"\[TBD[^\]]*\]")

# Patterns that denote a reporting-line phrase in the user's source message.
REPORTING_LINE_PATTERNS = [
    re.compile(r"\breports?\s+(?:up\s+)?to\s+([A-Z][A-Za-z\-']+(?:\s+[A-Z][A-Za-z\-']+)?)"),
    re.compile(r"\b([A-Z][A-Za-z\-']+(?:\s+[A-Z][A-Za-z\-']+)?)\s+is\s+\w+'s\s+manager\b"),
    re.compile(r"\brolls?\s+up\s+to\s+([A-Z][A-Za-z\-']+(?:\s+[A-Z][A-Za-z\-']+)?)"),
]


def build_reporting_line_banlist(source_text: str):
    """Names that appear in source ONLY inside a reporting-line phrase."""
    if not source_text:
        return set()
    candidates = set()
    rl_spans = []
    for pat in REPORTING_LINE_PATTERNS:
        for m in pat.finditer(source_text):
            candidates.add(m.group(1).strip())
            rl_spans.append(m.span())
    banned = set()
    for name in candidates:
        # Find all positions of name; if every position is inside an RL span, ban it.
        positions = []
        start = 0
        while True:
            i = source_text.find(name, start)
            if i == -1:
                break
            positions.append((i, i + len(name)))
            start = i + 1
        if not positions:
            continue
        all_in_rl = all(
            any(s <= p[0] and p[1] <= e for s, e in rl_spans) for p in positions
        )
        if all_in_rl:
            banned.add(name)
    return banned


def check_reporting_line_leaks(decisions, banlist):
    issues = []
    protected_fields = ("decision_maker", "decision", "rationale", "consequences")
    for d in decisions:
        did = d.get("id", "?")
        for field in protected_fields:
            val = d.get(field)
            if not isinstance(val, str):
                continue
            for name in banlist:
                if name in val:
                    issues.append(
                        f"⚠ {did}.{field} contains '{name}' — appears only as a reporting-line manager in source"
                    )
        for opt in d.get("options_considered") or []:
            opt_str = opt.get("option", "") if isinstance(opt, dict) else str(opt)
            for name in banlist:
                if name in opt_str:
                    issues.append(
                        f"⚠ {did}.options_considered contains '{name}' — appears only as a reporting-line manager in source"
                    )
    return issues


NAVY = "1F3A5F"
TBD_ORANGE = "B7471E"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TBD_FONT = Font(italic=True, color=TBD_ORANGE)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

STATUS_FILLS = {
    "Accepted": PatternFill("solid", fgColor="E2EFDA"),
    "Proposed": PatternFill("solid", fgColor="FFF2CC"),
    "Rejected": PatternFill("solid", fgColor="F8CBAD"),
    "Superseded": PatternFill("solid", fgColor="D9D9D9"),
    "Deferred": PatternFill("solid", fgColor="DDEBF7"),
}

VALID_STATUS = set(STATUS_FILLS.keys())

HEADERS = [
    ("ID", 8),
    ("Title", 36),
    ("Date", 12),
    ("Decision Maker", 22),
    ("Status", 12),
    ("Context", 40),
    ("Options Considered", 40),
    ("Decision", 40),
    ("Rationale", 40),
    ("Consequences", 36),
    ("Supersedes", 10),
    ("Review Date", 12),
]


def is_tbd(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return bool(TBD_PATTERN.search(v)) or not v.strip()
    return False


def format_options(options):
    if not options:
        return "[TBD — options not recorded]"
    parts = []
    for o in options:
        mark = "✓ " if o.get("chosen") else "  "
        parts.append(f"{mark}{o.get('option', '[TBD]')}")
    return "\n".join(parts)


def render(data, output_path: Path, source_text: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Decision Log"

    # Title row
    ws.cell(row=1, column=1, value=f"Decision Log — {data.get('header', {}).get('project_name', '[TBD]')}").font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    pm = data.get("header", {}).get("project_manager") or "[TBD]"
    compiled = data.get("header", {}).get("compiled_date") or "[TBD]"
    ws.cell(row=2, column=1, value=f"PM: {pm}   |   Compiled: {compiled}").font = Font(italic=True, size=10, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    # Header row
    header_row = 4
    for col, (label, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # Decision rows
    decisions = data.get("decisions") or []
    for i, d in enumerate(decisions):
        r = header_row + 1 + i
        values = [
            d.get("id", f"D-{i+1:03d}"),
            d.get("title", "[TBD]"),
            d.get("date", "[TBD]"),
            d.get("decision_maker", "[TBD]"),
            d.get("status", "Proposed"),
            d.get("context", ""),
            format_options(d.get("options_considered")),
            d.get("decision", "[TBD]"),
            d.get("rationale", "[TBD]"),
            d.get("consequences", "[TBD]"),
            d.get("supersedes_id", ""),
            d.get("review_date", ""),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if isinstance(v, str) and TBD_PATTERN.search(v):
                cell.font = TBD_FONT
            else:
                cell.font = Font(size=10, color="333333")
        # Status fill on status column
        status_cell = ws.cell(row=r, column=5)
        if d.get("status") in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[d.get("status")]
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            status_cell.font = Font(bold=True, size=10, color="333333")

    ws.row_dimensions[header_row].height = 24
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Integrity sheet: non-standard statuses, TBD decision makers, reporting-line leaks
    issues = []
    for d in decisions:
        if d.get("status") and d.get("status") not in VALID_STATUS:
            issues.append(f"{d.get('id', '?')} status '{d.get('status')}' is not standard")
        if is_tbd(d.get("decision_maker")) and d.get("status") == "Accepted":
            issues.append(f"{d.get('id', '?')} status Accepted but decision_maker is [TBD]")
    if source_text:
        banlist = build_reporting_line_banlist(source_text)
        issues.extend(check_reporting_line_leaks(decisions, banlist))
    if issues:
        ws2 = wb.create_sheet("Integrity")
        ws2.cell(row=1, column=1, value="Integrity Notes").font = Font(bold=True, size=12, color=NAVY)
        for i, issue in enumerate(issues, 2):
            ws2.cell(row=i, column=1, value=issue).font = Font(color=TBD_ORANGE, italic=True)
        ws2.column_dimensions["A"].width = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    gaps = []
    for d in decisions:
        did = d.get("id", "?")
        for f in ("decision_maker", "rationale"):
            if is_tbd(d.get(f)):
                gaps.append(f"{did}.{f}: [TBD]")
        if is_tbd(d.get("decision")) and d.get("status") not in ("Proposed", "Deferred"):
            gaps.append(f"{did}.decision: [TBD]")
    return gaps


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", help="Path to JSON input (omit to read stdin)")
    p.add_argument("--output", required=True)
    p.add_argument("--source", help="Path to user's raw message text; enables reporting-line leak detection")
    args = p.parse_args()

    source_text = ""
    if args.source:
        try:
            source_text = Path(args.source).read_text()
        except Exception as e:
            print(f"WARN: could not read --source file: {e}", file=sys.stderr)

    try:
        if args.input:
            data = json.loads(Path(args.input).read_text())
        else:
            data = json.loads(sys.stdin.read())
    except Exception as e:
        print(f"ERROR: could not parse input JSON: {e}", file=sys.stderr)
        return 2

    out = Path(args.output)
    gaps = render(data, out, source_text)
    print(f"Decision log written: {out}")
    print(f"TBD count: {len(gaps)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
