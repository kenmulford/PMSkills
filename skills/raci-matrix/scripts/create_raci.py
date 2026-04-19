#!/usr/bin/env python3
"""
Generate a RACI matrix .xlsx from a JSON input.

Enforces Project+ RACI rules (exactly one A per row, at least one R),
highlights TBDs, flags integrity issues on a dedicated sheet, and writes
a companion open_questions.md when there are many gaps.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TBD_PATTERN = re.compile(r"\[TBD[^\]]*\]")

# Colors
TBD_FONT_COLOR = "B7471E"    # burnt orange
HEADER_FILL = "1F3A5F"       # deep navy
HEADER_FONT_COLOR = "FFFFFF"
PHASE_FILL = "DCE6F1"        # light blue for phase group rows
ZEBRA_FILL = "F7F9FC"        # very light zebra

# RACI letter fills for quick visual scanning
LETTER_FILLS = {
    "R": "C6EFCE",   # green
    "A": "FFD966",   # amber
    "C": "BDD7EE",   # light blue
    "I": "E7E6E6",   # gray
    "R/A": "B7E1A1", # green-ish
}

VALID_LETTERS = {"R", "A", "C", "I", "R/A"}


def is_tbd_string(value) -> bool:
    return isinstance(value, str) and bool(TBD_PATTERN.search(value))


def thin_border() -> Border:
    side = Side(border_style="thin", color="B0B7C3")
    return Border(left=side, right=side, top=side, bottom=side)


def validate_raci(activities: list[dict], people: list[str]) -> list[str]:
    """Return a list of integrity issues. Empty list = clean.

    A row is only flagged if it has concrete assignments that violate the
    rules. Rows whose A or R is [TBD] are considered pending — the gap is
    surfaced on the Draft Gaps sheet, not as an integrity violation.
    """
    issues: list[str] = []
    for i, act in enumerate(activities, 1):
        name = act.get("name", f"Activity {i}")
        assignments = act.get("assignments", {}) or {}
        a_count = 0
        r_count = 0
        has_tbd = False
        for person, val in assignments.items():
            if not isinstance(val, str):
                continue
            if is_tbd_string(val):
                has_tbd = True
                continue
            v = val.strip().upper()
            if v == "A":
                a_count += 1
            elif v == "R":
                r_count += 1
            elif v == "R/A":
                a_count += 1
                r_count += 1
        # Only flag a missing A/R if the row has *some* concrete assignment;
        # fully-TBD rows are gaps, not integrity violations.
        row_has_concrete = (a_count + r_count) > 0 or any(
            isinstance(v, str) and v.strip().upper() in {"C", "I"}
            for v in assignments.values()
        )
        if row_has_concrete:
            if a_count == 0 and not has_tbd:
                issues.append(f"Row '{name}' has no Accountable (A). Every row needs exactly one A.")
            elif a_count > 1:
                issues.append(f"Row '{name}' has {a_count} Accountable assignments. Every row must have exactly one A.")
            if r_count == 0 and not has_tbd:
                issues.append(f"Row '{name}' has no Responsible (R). Every row needs at least one R.")

        for person in assignments.keys():
            if person not in people:
                issues.append(f"Row '{name}' references '{person}' who is not in the people list.")
    return issues


def collect_tbds(data: dict) -> list[str]:
    gaps: list[str] = []
    header = data.get("header", {}) or {}
    for k, v in header.items():
        if is_tbd_string(v) or (isinstance(v, str) and not v.strip()):
            gaps.append(f"header.{k}: {v or '[TBD — empty]'}")
    people = data.get("people") or []
    if not people:
        gaps.append("people: (empty — no team members provided)")
    for i, p in enumerate(people):
        if is_tbd_string(str(p)):
            gaps.append(f"people[{i}]: {p}")
    activities = data.get("activities") or []
    if not activities:
        gaps.append("activities: (empty — no activities provided)")
    for i, act in enumerate(activities):
        name = act.get("name", f"activity {i+1}")
        if is_tbd_string(name):
            gaps.append(f"activities[{i}].name: {name}")
        assignments = act.get("assignments", {}) or {}
        tbd_cells = [p for p, v in assignments.items() if is_tbd_string(str(v))]
        if tbd_cells:
            gaps.append(f"activities[{i}] '{name}' has TBD for: {', '.join(tbd_cells)}")
    return gaps


def style_cell_letter(cell, letter: str) -> None:
    letter_clean = letter.strip().upper()
    if letter_clean in LETTER_FILLS:
        cell.fill = PatternFill("solid", fgColor=LETTER_FILLS[letter_clean])
        cell.font = Font(bold=True, size=11)
    elif is_tbd_string(letter):
        cell.font = Font(color=TBD_FONT_COLOR, italic=True, size=10)
    else:
        cell.font = Font(size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()


def write_raci_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "RACI Matrix"

    header = data.get("header", {}) or {}
    project_name = header.get("project_name", "[TBD — no project name]")
    people = data.get("people") or []
    activities = data.get("activities") or []

    ncols = 1 + len(people)  # first col = activity name

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 2))
    tc = ws.cell(row=1, column=1, value=f"{project_name} — RACI Matrix")
    tc.font = Font(bold=True, size=16, color=HEADER_FILL)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Meta
    meta = [
        ("Project Manager", header.get("project_manager", "[TBD]")),
        ("Sponsor", header.get("sponsor", "[TBD]")),
        ("Date", header.get("date", "[TBD]")),
        ("Version", header.get("version", "v1")),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=max(ncols, 2))
        vc = ws.cell(row=i, column=2, value=str(v))
        vc.font = (Font(color=TBD_FONT_COLOR, italic=True, size=11)
                   if is_tbd_string(str(v)) else Font(size=11))

    # Legend row
    legend_row = 2 + len(meta) + 1
    legend_cell = ws.cell(row=legend_row, column=1,
                          value="Legend:  R=Responsible  A=Accountable (one per row)  C=Consulted  I=Informed")
    legend_cell.font = Font(italic=True, size=10, color="555555")
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=max(ncols, 2))

    # Matrix header row
    header_row = legend_row + 2
    activity_header = ws.cell(row=header_row, column=1, value="Activity / Deliverable")
    for i, person in enumerate(people, start=2):
        ws.cell(row=header_row, column=i, value=person)
    # Style header row
    hf = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    hfill = PatternFill("solid", fgColor=HEADER_FILL)
    for col in range(1, ncols + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[header_row].height = 42

    if not activities:
        row = header_row + 1
        ws.cell(row=row, column=1, value="[TBD — no activities provided]").font = Font(
            color=TBD_FONT_COLOR, italic=True
        )
    else:
        current_phase = None
        row = header_row
        for act in activities:
            phase = act.get("phase")
            if phase and phase != current_phase:
                row += 1
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
                pc = ws.cell(row=row, column=1, value=str(phase))
                pc.font = Font(bold=True, size=11, color=HEADER_FILL)
                pc.fill = PatternFill("solid", fgColor=PHASE_FILL)
                pc.alignment = Alignment(horizontal="left", vertical="center")
                current_phase = phase
            row += 1
            # Activity name cell
            name_cell = ws.cell(row=row, column=1, value=act.get("name", "[TBD]"))
            name_cell.alignment = Alignment(vertical="center", wrap_text=True)
            name_cell.border = thin_border()
            if is_tbd_string(name_cell.value or ""):
                name_cell.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
            else:
                name_cell.font = Font(size=11)
            # Zebra stripe
            if row % 2 == 0:
                name_cell.fill = PatternFill("solid", fgColor=ZEBRA_FILL)
            # Assignment cells
            assignments = act.get("assignments", {}) or {}
            for col_idx, person in enumerate(people, start=2):
                val = assignments.get(person, "")
                cell = ws.cell(row=row, column=col_idx, value=str(val) if val else "")
                style_cell_letter(cell, str(val) if val else "")
                if row % 2 == 0 and not (str(val).strip().upper() in LETTER_FILLS):
                    # zebra only if no letter fill
                    cell.fill = PatternFill("solid", fgColor=ZEBRA_FILL)

    # Column widths
    ws.column_dimensions["A"].width = 48
    for col_idx in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # Freeze header row + first column
    ws.freeze_panes = f"B{header_row + 1}"


def write_gaps_sheet(wb: Workbook, gaps: list[str]) -> None:
    ws = wb.create_sheet("Draft Gaps")
    ws.cell(row=1, column=1, value=f"Draft Gaps ({len(gaps)})").font = Font(
        bold=True, size=14, color=HEADER_FILL
    )
    intro = ws.cell(row=2, column=1, value=(
        "Unresolved items in this RACI draft. Resolve these with the sponsor "
        "before treating the matrix as final — every TBD is a potential source "
        "of accountability confusion on the project."
    ))
    intro.font = Font(italic=True, size=10)
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.row_dimensions[2].height = 42
    for i, g in enumerate(gaps, start=4):
        c = ws.cell(row=i, column=1, value=g)
        c.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30


def write_integrity_sheet(wb: Workbook, issues: list[str]) -> None:
    ws = wb.create_sheet("Integrity Issues")
    ws.cell(row=1, column=1, value=f"RACI Integrity Issues ({len(issues)})").font = Font(
        bold=True, size=14, color="B7471E"
    )
    intro = ws.cell(row=2, column=1, value=(
        "Project+ RACI rules: every row has EXACTLY ONE A (Accountable) and "
        "AT LEAST ONE R (Responsible). The issues below violate those rules "
        "and should be fixed before the matrix is used."
    ))
    intro.font = Font(italic=True, size=10)
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.row_dimensions[2].height = 42
    for i, issue in enumerate(issues, start=4):
        c = ws.cell(row=i, column=1, value=issue)
        c.font = Font(color="B7471E", size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    ws.column_dimensions["A"].width = 80


def render_raci(data: dict, output_path: Path) -> tuple[list[str], list[str]]:
    wb = Workbook()
    write_raci_sheet(wb, data)
    gaps = collect_tbds(data)
    issues = validate_raci(data.get("activities") or [], data.get("people") or [])
    if gaps:
        write_gaps_sheet(wb, gaps)
    if issues:
        write_integrity_sheet(wb, issues)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return gaps, issues


def write_open_questions(gaps: list[str], issues: list[str], output_path: Path) -> None:
    lines = [
        "# Open Questions for Sponsor / PM",
        "",
        f"This RACI draft has **{len(gaps)} unresolved items**"
        + (f" and **{len(issues)} integrity issues**" if issues else "")
        + ". Resolve these before using the matrix.",
        "",
        "## Draft Gaps",
        "",
    ]
    for g in gaps:
        lines.append(f"- {g}")
    if issues:
        lines.append("")
        lines.append("## Integrity Issues (Project+ rule violations)")
        lines.append("")
        for i in issues:
            lines.append(f"- {i}")
    lines.append("")
    output_path.write_text("\n".join(lines))


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a RACI matrix .xlsx")
    parser.add_argument("--input", help="Path to JSON input (omit to read stdin)")
    parser.add_argument("--output", required=True)
    parser.add_argument("--questions-threshold", type=int, default=5)
    args = parser.parse_args()

    output_path = Path(args.output)

    try:
        if args.input:
            data = json.loads(Path(args.input).read_text())
        else:
            data = json.loads(sys.stdin.read())
    except Exception as e:
        print(f"ERROR: could not parse input JSON: {e}", file=sys.stderr)
        return 2

    gaps, issues = render_raci(data, output_path)
    print(f"RACI written: {output_path}")
    print(f"TBD count: {len(gaps)}")
    if issues:
        print(f"Integrity issues: {len(issues)}")

    if len(gaps) >= args.questions_threshold or issues:
        oq_path = output_path.parent / "open_questions.md"
        write_open_questions(gaps, issues, oq_path)
        print(f"Open questions written: {oq_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
