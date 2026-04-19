#!/usr/bin/env python3
"""Generate a Project+ project budget .xlsx from JSON input.

Sheets:
- Budget: line items with category, amount, owner, status, notes
- Summary: totals by category, approved vs committed, variance, contingency
- Monthly Burn: if provided
- Draft Gaps, Integrity Issues, Notes
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


TBD_PATTERN = re.compile(r"\[TBD[^\]]*\]")

NAVY = "1F3A5F"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
TBD_ORANGE = "B7471E"
TBD_FILL = "FCE8D4"
GOOD_GREEN = "C6E0B4"
BAD_RED = "F4B4B4"

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

VALID_CATEGORIES = {"Labor", "Software", "Vendor", "Hardware", "Travel", "Contingency", "Other"}
VALID_STATUS = {"Committed", "Planned", "Contracted", "Unfunded", "TBD"}


def is_tbd(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return bool(TBD_PATTERN.search(v)) or not v.strip()
    return False


def _header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _tbd_style(cell):
    cell.font = Font(italic=True, color=TBD_ORANGE)
    cell.fill = PatternFill("solid", fgColor=TBD_FILL)


def _money(n):
    if isinstance(n, (int, float)):
        return f"${n:,.0f}"
    return n


def compute_totals(line_items, approved_total, contingency_percent):
    """Return (committed_sum, contingency_amount, totals_by_category, any_tbd_amount)."""
    totals_by_cat = {}
    committed = 0
    any_tbd = False
    for li in line_items:
        amt = li.get("amount")
        cat = li.get("category", "Other")
        if is_tbd(amt) or not isinstance(amt, (int, float)):
            any_tbd = True
            totals_by_cat.setdefault(cat, {"amount": 0, "tbd": True})
            totals_by_cat[cat]["tbd"] = True
            continue
        committed += amt
        totals_by_cat.setdefault(cat, {"amount": 0, "tbd": False})
        totals_by_cat[cat]["amount"] += amt

    contingency_amount = None
    if (isinstance(contingency_percent, (int, float))
            and isinstance(approved_total, (int, float))):
        contingency_amount = approved_total * (contingency_percent / 100.0)

    return committed, contingency_amount, totals_by_cat, any_tbd


def render_budget_sheet(wb, data):
    ws = wb.active
    ws.title = "Budget"
    h = data.get("header", {}) or {}

    ws["A1"] = "Project:"
    ws["B1"] = h.get("project_name", "[TBD]")
    ws["A2"] = "PM:"
    ws["B2"] = h.get("project_manager", "[TBD]")
    ws["D2"] = "Sponsor:"
    ws["E2"] = h.get("sponsor", "[TBD]")
    ws["A3"] = "Fiscal Year:"
    ws["B3"] = h.get("fiscal_year", "[TBD]")
    ws["D3"] = "Approved Total:"
    at = h.get("approved_total", "[TBD]")
    ws["E3"] = _money(at) if isinstance(at, (int, float)) else at
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=4).font = Font(bold=True)

    headers = ["ID", "Category", "Description", "Amount", "Vendor", "Owner", "Status", "Notes"]
    start = 5
    _header_row(ws, headers, row=start)

    for i, li in enumerate(data.get("line_items") or []):
        r = start + 1 + i
        amt = li.get("amount")
        vals = [
            li.get("id", f"L{i+1:03d}"),
            li.get("category", "[TBD]"),
            li.get("description", "[TBD]"),
            _money(amt) if isinstance(amt, (int, float)) else amt,
            li.get("vendor") or "",
            li.get("owner", "[TBD]"),
            li.get("status") or "Planned",
            li.get("notes") or "",
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if isinstance(v, str) and TBD_PATTERN.search(v):
                _tbd_style(c)
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right")

    widths = [8, 14, 40, 14, 20, 18, 14, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"B{start + 1}"


def render_summary(wb, data):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Budget Summary"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)

    h = data.get("header", {}) or {}
    approved = h.get("approved_total", "[TBD]")
    contingency_pct = h.get("contingency_percent", "[TBD]")
    line_items = data.get("line_items") or []

    committed, contingency_amt, by_cat, any_tbd = compute_totals(
        line_items, approved, contingency_pct
    )

    _header_row(ws, ["Category", "Committed", "Status"], row=3)
    row = 4
    for cat in sorted(by_cat.keys()):
        info = by_cat[cat]
        ws.cell(row=row, column=1, value=cat).border = BORDER
        ws.cell(row=row, column=2, value=_money(info["amount"])).border = BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        status = "includes TBD items" if info["tbd"] else "fully specified"
        c = ws.cell(row=row, column=3, value=status)
        c.border = BORDER
        if info["tbd"]:
            _tbd_style(c)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Committed subtotal").font = Font(bold=True)
    ws.cell(row=row, column=2, value=_money(committed)).font = Font(bold=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    row += 1

    ws.cell(row=row, column=1, value="Contingency")
    if contingency_amt is not None:
        ws.cell(row=row, column=2, value=_money(contingency_amt))
        ws.cell(row=row, column=3,
                value=f"{contingency_pct}% of approved {_money(approved)}")
    else:
        c = ws.cell(row=row, column=2, value="[TBD — no percent set]")
        _tbd_style(c)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    row += 1

    ws.cell(row=row, column=1, value="Approved total").font = Font(bold=True)
    ws.cell(row=row, column=2,
            value=_money(approved) if isinstance(approved, (int, float)) else approved
            ).font = Font(bold=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    row += 1

    ws.cell(row=row, column=1, value="Variance (approved − committed − contingency)").font = Font(bold=True)
    if (isinstance(approved, (int, float)) and not any_tbd
            and contingency_amt is not None):
        var = approved - committed - contingency_amt
        c = ws.cell(row=row, column=2, value=_money(var))
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=GOOD_GREEN if var >= 0 else BAD_RED)
    else:
        reason = []
        if not isinstance(approved, (int, float)):
            reason.append("no approved baseline")
        if any_tbd:
            reason.append("TBD line items")
        if contingency_amt is None:
            reason.append("contingency TBD")
        c = ws.cell(row=row, column=2, value=f"[TBD — {', '.join(reason)}]")
        _tbd_style(c)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

    for i, w in enumerate([35, 20, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def render_burn(wb, data):
    burn = data.get("monthly_burn") or []
    if not burn:
        return
    ws = wb.create_sheet("Monthly Burn")
    ws["A1"] = "Monthly Burn"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    _header_row(ws, ["Month", "Amount"], row=3)
    total = 0
    for i, b in enumerate(burn, 4):
        ws.cell(row=i, column=1, value=b.get("month", "[TBD]")).border = BORDER
        amt = b.get("amount")
        if isinstance(amt, (int, float)):
            total += amt
            c = ws.cell(row=i, column=2, value=_money(amt))
        else:
            c = ws.cell(row=i, column=2, value=amt or "[TBD]")
            _tbd_style(c)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right")
    ws.cell(row=i + 2, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=i + 2, column=2, value=_money(total)).font = Font(bold=True)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18


def collect_gaps(data):
    gaps = []
    h = data.get("header", {}) or {}
    for k in ("project_name", "project_manager", "sponsor", "approved_total", "contingency_percent"):
        if is_tbd(h.get(k)):
            gaps.append(f"header.{k}: {h.get(k) or '[TBD — empty]'}")
    items = data.get("line_items") or []
    if not items:
        gaps.append("line_items: (empty)")
    for li in items:
        lid = li.get("id", "?")
        for f in ("category", "description", "amount"):
            if is_tbd(li.get(f)):
                gaps.append(f"{lid}.{f}: {li.get(f) or '[TBD — empty]'}")
    return gaps


def render_gaps(wb, gaps):
    ws = wb.create_sheet("Draft Gaps")
    ws["A1"] = f"Draft Gaps ({len(gaps)})"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    if not gaps:
        ws["A3"] = "None — budget is fully specified."
        return
    for i, g in enumerate(gaps, 1):
        c = ws.cell(row=2 + i, column=1, value=g)
        _tbd_style(c)
        c.border = BORDER
    ws.column_dimensions["A"].width = 100


def render_integrity(wb, data):
    ws = wb.create_sheet("Integrity Issues")
    ws["A1"] = "Integrity Issues"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)

    issues = []
    line_items = data.get("line_items") or []
    for li in line_items:
        lid = li.get("id", "?")
        cat = li.get("category")
        if not is_tbd(cat) and cat not in VALID_CATEGORIES:
            issues.append(f"{lid}: non-standard category '{cat}' — expected one of {sorted(VALID_CATEGORIES)}")
        status = li.get("status")
        if status and status not in VALID_STATUS:
            issues.append(f"{lid}: non-standard status '{status}'")
        desc = (li.get("description") or "").lower()
        if "unallocated" in desc or "rounding" in desc or "plug" in desc:
            issues.append(f"{lid}: looks like a reconciling plug line — remove and let Variance tell the truth")

    h = data.get("header", {}) or {}
    committed, contingency_amt, _, any_tbd = compute_totals(
        line_items, h.get("approved_total"), h.get("contingency_percent")
    )
    approved = h.get("approved_total")
    if (isinstance(approved, (int, float)) and contingency_amt is not None
            and not any_tbd and committed + contingency_amt > approved):
        issues.append(
            f"Committed {_money(committed)} + contingency {_money(contingency_amt)} "
            f"exceeds approved {_money(approved)}"
        )

    if not issues:
        ws["A3"] = "No integrity issues detected."
        return
    for i, item in enumerate(issues, 1):
        c = ws.cell(row=2 + i, column=1, value=item)
        _tbd_style(c)
    ws.column_dimensions["A"].width = 100


def render_notes(wb):
    ws = wb.create_sheet("Notes")
    ws["A1"] = "Budget Conventions"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    notes = [
        "• Amounts are in project currency (assumed USD unless noted).",
        "• Approved total is the baseline frozen at sponsor approval; changes go through change control.",
        "• Variance = Approved − Committed − Contingency.",
        "• Contingency reserve covers known risks; management reserve (unknown-unknowns) is not in this baseline.",
        "• [TBD — reason] cells are honest gaps: the source user message did not contain that information.",
        "• Rows labeled 'Unallocated' or 'Rounding' are flagged as reconciling plug lines — they should not exist.",
    ]
    for i, n in enumerate(notes, 3):
        ws.cell(row=i, column=1, value=n)
    ws.column_dimensions["A"].width = 110


def write_open_questions(gaps, output):
    lines = ["# Open Questions — Project Budget", "",
             f"This draft has **{len(gaps)} unresolved items**. Resolve before presenting to the sponsor.", ""]
    for g in gaps:
        lines.append(f"- {g}")
    output.write_text("\n".join(lines))


def render(data, output_path: Path, questions_threshold=5):
    wb = Workbook()
    render_budget_sheet(wb, data)
    render_summary(wb, data)
    render_burn(wb, data)
    gaps = collect_gaps(data)
    render_gaps(wb, gaps)
    render_integrity(wb, data)
    render_notes(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    if len(gaps) >= questions_threshold:
        oq = output_path.parent / "open_questions.md"
        write_open_questions(gaps, oq)
        return gaps, oq
    return gaps, None


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", help="Path to JSON input (omit to read stdin)")
    p.add_argument("--output", required=True)
    p.add_argument("--questions-threshold", type=int, default=5)
    args = p.parse_args()

    try:
        if args.input:
            data = json.loads(Path(args.input).read_text())
        else:
            data = json.loads(sys.stdin.read())
    except Exception as e:
        print(f"ERROR: could not parse input JSON: {e}", file=sys.stderr)
        return 2

    out = Path(args.output)
    gaps, oq = render(data, out, args.questions_threshold)
    print(f"Budget written: {out}")
    print(f"TBD count: {len(gaps)}")
    if oq:
        print(f"Open questions: {oq}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
