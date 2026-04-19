#!/usr/bin/env python3
"""
Generate a project team roster .xlsx from a JSON input.

Enforces CompTIA Project+ team roster columns, highlights TBDs in burnt
orange, emits a "Draft Gaps" sheet listing every unresolved field, and
writes a companion open_questions.md when the TBD count is large.

The skill calls this script instead of writing openpyxl code from scratch.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TBD_PATTERN = re.compile(r"\[TBD[^\]]*\]")

# Colors
TBD_FONT_COLOR = "B7471E"  # burnt orange
HEADER_FILL = "1F3A5F"     # deep navy
HEADER_FONT_COLOR = "FFFFFF"
ZEBRA_FILL = "F2F4F7"      # light gray zebra


COLUMNS = [
    ("name", "Name"),
    ("job_position", "Job Position"),
    ("core_or_extended", "Core / Extended"),
    ("subteam", "Subteam / Grouping"),
    ("responsibilities", "Responsibilities"),
    ("availability", "Availability"),
    ("manager", "Manager"),
    ("organization", "Organization"),
    ("contact_method", "Preferred Contact"),
]


def is_tbd_string(value) -> bool:
    return isinstance(value, str) and bool(TBD_PATTERN.search(value))


def collect_tbds(data: dict) -> list[str]:
    """Walk the schema and return human-readable TBD locations."""
    gaps: list[str] = []

    header = data.get("header", {}) or {}
    for k, v in header.items():
        if is_tbd_string(v) or (isinstance(v, str) and not v.strip()):
            gaps.append(f"header.{k}: {v}")

    members = data.get("members") or []
    if not members:
        gaps.append("members: (empty — no team members provided)")
    for i, m in enumerate(members):
        for key, _label in COLUMNS:
            val = m.get(key, "")
            if is_tbd_string(val) or (isinstance(val, str) and not val.strip()):
                name = m.get("name") or f"row {i+1}"
                gaps.append(f"members[{i}] ({name}).{key}: {val or '[TBD — empty]'}")
    return gaps


def style_header_row(ws, ncols: int) -> None:
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    center = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 22


def apply_cell_style(cell, value, zebra: bool) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if is_tbd_string(str(value)):
        cell.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
    else:
        cell.font = Font(size=11)
    if zebra:
        cell.fill = PatternFill("solid", fgColor=ZEBRA_FILL)


def autosize_columns(ws, col_widths: dict) -> None:
    for idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_roster_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "Team Roster"

    header = data.get("header", {}) or {}
    project_name = header.get("project_name", "[TBD — no project name]")

    # Title row (merged)
    ncols = len(COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value=f"{project_name} — Project Team Roster")
    title_cell.font = Font(bold=True, size=16, color=HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Header metadata block
    meta = [
        ("Project Manager", header.get("project_manager", "[TBD]")),
        ("Sponsor", header.get("sponsor", "[TBD]")),
        ("Date", header.get("date", "[TBD]")),
        ("Version", header.get("version", "v1")),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=ncols)
        vc = ws.cell(row=i, column=2, value=str(v))
        if is_tbd_string(str(v)):
            vc.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
        else:
            vc.font = Font(size=11)

    # Column headers (after meta block + 1 spacer row)
    header_row = 2 + len(meta) + 1  # 7
    for col_idx, (_key, label) in enumerate(COLUMNS, start=1):
        ws.cell(row=header_row, column=col_idx, value=label)
    # Style header row — reuse helper but targeted at header_row
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for col in range(1, ncols + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 22

    # Member rows
    members = data.get("members") or []
    if not members:
        members = [{
            "name": "[TBD — no team members provided]",
            "job_position": "[TBD]",
            "core_or_extended": "[TBD]",
            "subteam": "[TBD]",
            "responsibilities": "[TBD]",
            "availability": "[TBD]",
            "manager": "[TBD]",
            "organization": "[TBD]",
            "contact_method": "[TBD]",
        }]

    for r_offset, m in enumerate(members, start=1):
        row = header_row + r_offset
        zebra = (r_offset % 2 == 0)
        for col_idx, (key, _label) in enumerate(COLUMNS, start=1):
            val = m.get(key, "")
            if val is None or (isinstance(val, str) and not val.strip()):
                val = "[TBD — empty]"
            cell = ws.cell(row=row, column=col_idx, value=str(val))
            apply_cell_style(cell, val, zebra)

    # Column widths tuned for roster readability
    widths = {
        1: 22,  # Name
        2: 26,  # Job Position
        3: 14,  # Core/Extended
        4: 18,  # Subteam
        5: 40,  # Responsibilities
        6: 16,  # Availability
        7: 20,  # Manager
        8: 20,  # Organization
        9: 22,  # Contact
    }
    autosize_columns(ws, widths)

    # Freeze the member-table header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_gaps_sheet(wb: Workbook, gaps: list[str]) -> None:
    ws = wb.create_sheet("Draft Gaps")
    ws.cell(row=1, column=1, value=f"Draft Gaps ({len(gaps)})").font = Font(
        bold=True, size=14, color=HEADER_FILL
    )
    intro = (
        "This roster draft has unresolved items that need sponsor/PM input "
        "before it should be treated as final. They are listed below and "
        "highlighted in orange on the Team Roster sheet."
    )
    ic = ws.cell(row=2, column=1, value=intro)
    ic.alignment = Alignment(wrap_text=True, vertical="top")
    ic.font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.row_dimensions[2].height = 40

    header_cell = ws.cell(row=4, column=1, value="Field")
    header_cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    detail_cell = ws.cell(row=4, column=2, value="Detail")
    detail_cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    detail_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)

    for i, g in enumerate(gaps, start=5):
        # split "field: value" into two columns if possible
        if ": " in g:
            field, detail = g.split(": ", 1)
        else:
            field, detail = g, ""
        fc = ws.cell(row=i, column=1, value=field)
        dc = ws.cell(row=i, column=2, value=detail)
        for c in (fc, dc):
            c.font = Font(color=TBD_FONT_COLOR, italic=True, size=11)
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 60


def render_roster(data: dict, output_path: Path) -> list[str]:
    wb = Workbook()
    write_roster_sheet(wb, data)
    gaps = collect_tbds(data)
    if gaps:
        write_gaps_sheet(wb, gaps)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return gaps


def write_open_questions(gaps: list[str], output_path: Path) -> None:
    lines = [
        "# Open Questions for Sponsor / PM",
        "",
        f"This roster draft has **{len(gaps)} unresolved items**. Resolve these before treating the roster as final.",
        "",
    ]
    by_section: dict[str, list[str]] = {}
    for g in gaps:
        section = g.split(".")[0].split("[")[0].strip() or "general"
        by_section.setdefault(section, []).append(g)
    for section, items in sorted(by_section.items()):
        lines.append(f"## {section}")
        lines.append("")
        for item in items:
            lines.append(f"- {item}")
        lines.append("")
    output_path.write_text("\n".join(lines))


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a project team roster .xlsx")
    parser.add_argument("--input", help="Path to JSON input file (omit to read from stdin)")
    parser.add_argument("--output", required=True, help="Path to output .xlsx file")
    parser.add_argument(
        "--questions-threshold",
        type=int,
        default=5,
        help="Emit open_questions.md when TBD count >= this (default 5)",
    )
    args = parser.parse_args()

    output_path = Path(args.output)

    try:
        if args.input:
            data = json.loads(Path(args.input).read_text())
        else:
            data = json.loads(sys.stdin.read())
    except Exception as e:
        print(f"ERROR: could not parse input JSON: {e}", file=sys.stderr)
        return 2

    gaps = render_roster(data, output_path)
    print(f"Roster written: {output_path}")
    print(f"TBD count: {len(gaps)}")

    if len(gaps) >= args.questions_threshold:
        oq_path = output_path.parent / "open_questions.md"
        write_open_questions(gaps, oq_path)
        print(f"Open questions written: {oq_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
